"""
REPORT_OUTPUT — terminal node. Writes the final Excel artifact.

Layout (fixed order):

  1. Cover sheet — trader, instrument, alert date, disposition badge.
  2. Executive Summary — `ctx.executive_summary` from CONSOLIDATED_SUMMARY.
  3. Section Summaries — one block per `ctx.sections[name]`.
  4. Data tabs — declared in `config.tabs`.

Two flavours of data tab:

  • Static  — `{name, dataset, include_highlights}`. Picks a single
    DataFrame out of `ctx.datasets`. When `include_highlights` is
    true and a sibling `<dataset>_highlighted` exists, we use that
    one (DATA_HIGHLIGHTER's output) so coloured rows render.
  • Dynamic — `{expand_from, as, name, dataset, include_highlights}`.
    `expand_from` resolves a ref expression to an iterable
    (context list, MAP results dict, dataset column …) and emits one
    tab per item. Both `name` and `dataset` are templates with the
    `as` slot bound to the current item — that's how the FXFRO
    workflow gets a tab per book and FISL gets a tab per venue.

Topology rule (validator-enforced): there is at least one
REPORT_OUTPUT, it has no outgoing edges (it's terminal).

Cloud Run note: the container filesystem is read-only except for
`/tmp`, so we honour `DBSHERPA_OUTPUT_DIR` to redirect writes without
touching workflow JSON.
"""
import os
import re
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..context import RunContext
from ..node_spec import NodeSpec, _spec_from_yaml
from ..prompt_context import SafeMap as _SafeMap
from ..refs import ResolveError, resolve_ref, resolve_template

# Single-ref matcher (full-string). Different shape from refs.REF_RE which
# matches refs anywhere in a template — here we accept ONLY a whole-string
# `{ref}` with optional surrounding whitespace.
_SINGLE_REF_RE = re.compile(r"^\s*\{([^}]+)\}\s*$")


def _iter_for_expand(expr: str, ctx: RunContext):
    """Resolve an `expand_from` expression to an iterable of items.

    Accepts a single ref expression (`{...}`) that resolves to:
      • list / tuple                → iterated as-is
      • dict                        → iterated by keys
      • pandas.Series               → unique values, dropna
      • MAP result dict {results:.} → result keys
      • anything truthy & iterable  → iter()
    Returns [] for unresolvable refs (silent skip — keeps reports
    rendering when an upstream node didn't produce its expected list).
    """
    m = _SINGLE_REF_RE.match(expr or "")
    if not m:
        return []
    try:
        value = resolve_ref(m.group(1), ctx)
    except ResolveError:
        return []
    if value is None:
        return []
    if isinstance(value, dict):
        if "results" in value and isinstance(value["results"], dict):
            return list(value["results"].keys())
        return list(value.keys())
    if isinstance(value, pd.Series):
        return list(value.dropna().unique())
    if isinstance(value, pd.DataFrame):
        return list(range(len(value)))
    if isinstance(value, (list, tuple)):
        return list(value)
    return [value]


def _hex_fill(hex_colour: str) -> PatternFill:
    colour = hex_colour.lstrip("#").upper().zfill(6)
    return PatternFill(fill_type="solid", fgColor=colour)


def _write_df(ws, df: pd.DataFrame, freeze: bool = True) -> None:
    header_fill = _hex_fill("1C2333")
    header_font = Font(bold=True, color="F9FAFB", size=10)

    for ci, col in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=ci, value=str(col))
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

    if freeze:
        ws.freeze_panes = "A2"

    colour_col_idx = (
        df.columns.get_loc("_highlight_colour") + 1
        if "_highlight_colour" in df.columns
        else None
    )

    for ri, row_vals in enumerate(df.itertuples(index=False), 2):
        row_colour: str | None = None
        if colour_col_idx is not None:
            c = row_vals[colour_col_idx - 1]
            if c and c not in ("#FFFFFF", ""):
                row_colour = c

        for ci, val in enumerate(row_vals, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            if row_colour:
                cell.fill = _hex_fill(row_colour)
                cell.font = Font(size=9, color="000000")
            else:
                cell.font = Font(size=9)

    # Auto-width (capped at 40)
    for ci, col in enumerate(df.columns, 1):
        max_len = len(str(col))
        if len(df) > 0:
            max_len = max(max_len, df[col].astype(str).str.len().max())
        ws.column_dimensions[get_column_letter(ci)].width = min(int(max_len) + 2, 40)


def _df_for_excel(df: pd.DataFrame) -> pd.DataFrame:
    """Stringify datetime columns; Excel can't hold timezone-aware datetimes."""
    df = df.copy()
    for col in df.select_dtypes(include=["datetime64[ns, UTC]", "datetime64[ns]", "datetimetz"]).columns:
        df[col] = df[col].astype(str)
    # Lists/dicts in cells break Excel
    for col in df.columns:
        if df[col].apply(lambda x: isinstance(x, (list, dict))).any():
            df[col] = df[col].apply(str)
    return df


def handle_report_output(node: dict, ctx: RunContext) -> None:
    cfg = node.get("config", {})
    raw_output_path: str = cfg.get("output_path", "output/report.xlsx")
    # Resolve {context.xxx} and strip any remaining {unresolved} placeholders so filenames are safe.
    output_path = resolve_template(raw_output_path, ctx)
    output_path = re.sub(r"\{[^}]*\}", "", output_path)

    # On Cloud Run the repo-relative "output/" directory is not
    # writable — the container filesystem is read-only except for
    # `/tmp`. We honour `DBSHERPA_OUTPUT_DIR` so deployments can
    # redirect report writes to `/tmp/output` (ephemeral demo) or a
    # GCS FUSE mount (persistent) without touching workflow YAML.
    output_root = os.environ.get("DBSHERPA_OUTPUT_DIR")
    if output_root and not os.path.isabs(output_path):
        # Normalise by stripping a leading "output/" so the env var
        # is authoritative (avoids `/tmp/output/output/report.xlsx`).
        norm = output_path
        if norm.startswith("output/"):
            norm = norm[len("output/"):]
        output_path = os.path.join(output_root, norm)
    raw_tabs: list[dict] = list(cfg.get("tabs") or [])

    # Expand any tab declaring `expand_from` into one tab per item in
    # the resolved iterable. The iterable can be:
    #   • a context list (`{context.book_list}`)
    #   • a column of a dataset (`{ladder.bucket}` → unique values)
    #   • a MAP result dict (`{per_book.results}` → keys)
    # Each item is bound under the configured `as` name (default "item")
    # and substituted into both `name` and `dataset` template strings.
    tabs: list[dict] = []
    for tab in raw_tabs:
        expand_expr = tab.get("expand_from")
        if not expand_expr:
            tabs.append(tab)
            continue
        items = _iter_for_expand(expand_expr, ctx)
        bind = tab.get("as", "item")
        for item in items:
            slot = {bind: item}
            ds_name = (tab.get("dataset") or "").format_map(_SafeMap(slot))
            if ds_name not in ctx.datasets:
                continue
            tabs.append({
                "name": (tab.get("name", str(item))).format_map(_SafeMap(slot))[:31],
                "dataset": ds_name,
                "include_highlights": bool(tab.get("include_highlights", True)),
            })

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    def build_cover() -> None:
        ws_cover = wb.create_sheet("Cover")
        ws_cover.sheet_view.showGridLines = False
        bg = _hex_fill("0D1B2A")
        for row in ws_cover.iter_rows(min_row=1, max_row=40, min_col=1, max_col=15):
            for cell in row:
                cell.fill = bg

        def cover_cell(addr: str, text: str, bold: bool = False, size: int = 12, color: str = "F9FAFB"):
            c = ws_cover[addr]
            c.value = text
            c.font = Font(bold=bold, size=size, color=color)

        cover_cell("B2", "dbSherpa — Trade Surveillance Report", bold=True, size=22, color="F59E0B")
        cover_cell("B4", f"Trader:       {ctx.get('trader_id', 'N/A')}", size=13)
        cover_cell("B5", f"Instrument:   {ctx.get('currency_pair', 'N/A')}", size=13)
        cover_cell("B6", f"Alert Date:   {ctx.get('alert_date', 'N/A')}", size=13)
        disp_colour = {"ESCALATE": "EF4444", "REVIEW": "F59E0B", "DISMISS": "10B981"}.get(ctx.disposition, "F9FAFB")
        cover_cell("B8", f"Disposition:  {ctx.disposition}", bold=True, size=16, color=disp_colour)
        cover_cell("B9", f"Signal Flags: {ctx.get('flag_count', 0)}", size=13)
        ws_cover.column_dimensions["B"].width = 60

    def build_exec_summary() -> None:
        if not ctx.executive_summary:
            return
        ws_exec = wb.create_sheet("Executive Summary")
        ws_exec["A1"] = "EXECUTIVE SUMMARY"
        ws_exec["A1"].font = Font(bold=True, size=14, color="0D1B2A")
        ws_exec["A3"] = ctx.executive_summary
        ws_exec["A3"].alignment = Alignment(wrap_text=True, vertical="top")
        ws_exec.column_dimensions["A"].width = 120
        ws_exec.row_dimensions[3].height = 500

    def build_section_summaries() -> None:
        if not ctx.sections:
            return
        ws_sec = wb.create_sheet("Section Summaries")
        row = 1
        for name, sec in ctx.sections.items():
            ws_sec.cell(row=row, column=1, value=name.upper().replace("_", " ")).font = Font(bold=True, size=12)
            row += 1
            cell = ws_sec.cell(row=row, column=1, value=sec.get("narrative", ""))
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws_sec.row_dimensions[row].height = 200
            row += 3
        ws_sec.column_dimensions["A"].width = 110

    def build_summary_sheets() -> None:
        build_cover()
        build_exec_summary()
        build_section_summaries()

    def build_data_tabs() -> None:
        local_tabs = tabs or [
            {"name": ds_name[:31], "dataset": ds_name, "include_highlights": True}
            for ds_name in ctx.datasets
            if not ds_name.endswith("_highlighted")
        ]
        for tab in local_tabs:
            tab_name: str = tab.get("name", "Data")[:31]
            ds_name: str = tab.get("dataset", "")
            use_highlights: bool = tab.get("include_highlights", True)

            df = ctx.datasets.get(ds_name)
            if df is None:
                continue

            highlighted_key = f"{ds_name}_highlighted"
            if use_highlights and highlighted_key in ctx.datasets:
                df = ctx.datasets[highlighted_key]

            ws = wb.create_sheet(tab_name)
            _write_df(ws, _df_for_excel(df))

    build_summary_sheets()
    build_data_tabs()

    wb.save(output_path)
    ctx.report_path = output_path
    ctx.set("report_path", output_path)


NODE_SPEC: NodeSpec = _spec_from_yaml(Path(__file__).with_suffix(".yaml"), handle_report_output)
